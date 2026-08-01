from __future__ import annotations

import csv
import re
from pathlib import Path

from antismurf.models.player import parse_handle_parts
from antismurf.roster.models import PlayerRosterEntry

DISPLAY_NAME_HEADERS = {"玩家名", "display_name", "昵称", "name", "玩家名称"}
HANDLE_HEADERS = {"句柄", "handle", "句柄名"}
REMARK_HEADERS = {"备注", "remark", "note", "notes", "说明"}


def normalize_display_name(name: str) -> str:
    text = name.strip()
    match = re.match(r"(?:<#[^<>]{0,32}>)?#?(.+)$", text)
    if match:
        return match.group(1).strip()
    return text


def parse_roster_rows(rows: list[dict[str, str]], *, source: str = "import") -> list[PlayerRosterEntry]:
    entries: list[PlayerRosterEntry] = []
    for row in rows:
        display_name = str(row.get("display_name", "")).strip()
        handle = str(row.get("handle", "")).strip()
        remark = str(row.get("remark", "")).strip()
        if not handle and not display_name:
            continue
        if handle and parse_handle_parts(handle) is None:
            continue
        entries.append(
            PlayerRosterEntry(
                handle=handle,
                display_name=display_name,
                remark=remark,
                source=source,
            )
        )
    return entries


def _map_header(header: str) -> str | None:
    key = header.strip().lower()
    if key in {h.lower() for h in DISPLAY_NAME_HEADERS}:
        return "display_name"
    if key in {h.lower() for h in HANDLE_HEADERS}:
        return "handle"
    if key in {h.lower() for h in REMARK_HEADERS}:
        return "remark"
    return None


def _rows_from_header_dicts(headers: list[str], data_rows: list[list[str]]) -> list[dict[str, str]]:
    mapping: dict[int, str] = {}
    for index, header in enumerate(headers):
        field_name = _map_header(header)
        if field_name:
            mapping[index] = field_name

    if "handle" not in mapping.values():
        return []

    result: list[dict[str, str]] = []
    for row in data_rows:
        item: dict[str, str] = {"display_name": "", "handle": "", "remark": ""}
        for index, field_name in mapping.items():
            if index < len(row):
                item[field_name] = str(row[index] or "").strip()
        if item["handle"] or item["display_name"]:
            result.append(item)
    return result


def parse_csv_file(path: Path) -> list[PlayerRosterEntry]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return []
    headers = rows[0]
    data_rows = rows[1:]
    mapped = _rows_from_header_dicts(headers, data_rows)
    return parse_roster_rows(mapped, source="local_file")


def parse_xlsx_file(path: Path) -> list[PlayerRosterEntry]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):
        rows.append([str(cell).strip() if cell is not None else "" for cell in row])
    workbook.close()
    if not rows:
        return []
    mapped = _rows_from_header_dicts(rows[0], rows[1:])
    return parse_roster_rows(mapped, source="local_file")


def parse_roster_file(path: str | Path) -> list[PlayerRosterEntry]:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return parse_csv_file(file_path)
    if suffix in (".xlsx", ".xlsm"):
        return parse_xlsx_file(file_path)
    raise ValueError(f"不支持的名册文件格式: {suffix}")


def parse_roster_bytes(content: bytes, *, hint: str = "") -> list[dict[str, str]]:
    """Parse roster rows from CSV/XLSX bytes (remote export)."""
    suffix = Path(hint.split("?", 1)[0]).suffix.lower()
    if suffix == ".csv" or content.lstrip().startswith(b"\xef\xbb\xbf") or b"," in content[:200]:
        text = content.decode("utf-8-sig", errors="replace")
        import csv

        rows = list(csv.reader(text.splitlines()))
        if not rows:
            return []
        mapped = _rows_from_header_dicts(rows[0], rows[1:])
        return [
            {
                "display_name": e.display_name,
                "handle": e.handle,
                "remark": e.remark,
            }
            for e in parse_roster_rows(mapped, source="http_url")
        ]
    if suffix in (".xlsx", ".xlsm", ".xls") or content[:2] == b"PK":
        import io

        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        sheet_rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            sheet_rows.append([str(cell).strip() if cell is not None else "" for cell in row])
        workbook.close()
        if not sheet_rows:
            return []
        mapped = _rows_from_header_dicts(sheet_rows[0], sheet_rows[1:])
        return [
            {
                "display_name": e.display_name,
                "handle": e.handle,
                "remark": e.remark,
            }
            for e in parse_roster_rows(mapped, source="http_url")
        ]
    raise ValueError("无法识别远程名册格式，请使用 CSV 或 XLSX 导出链接")
